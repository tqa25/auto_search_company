import openpyxl

from src.excel_handler import ExcelWriter


class FakeDb:
    def get_all_companies(self):
        return [
            {
                "id": 1,
                "original_name": "Company A",
                "vietnamese_name": "Cong ty A",
                "tax_code": "0101",
                "status": "done",
                "business_status": "Tạm ngừng KD có thời hạn",
                "business_status_category": "inactive_stop",
                "business_status_source_url": "https://masothue.com/a",
            },
            {
                "id": 2,
                "original_name": "Company B",
                "vietnamese_name": "Cong ty B",
                "tax_code": "0202",
                "status": "done",
                "business_status": "Đang hoạt động",
                "business_status_category": "active",
                "business_status_source_url": "https://masothue.com/b",
            },
        ]

    def get_gemini_quick_results_for_company(self, company_id):
        if company_id == 1:
            return [{"phone": "+84 291 3957555", "email": "a@example.com", "address": "Addr A", "website": "a.vn"}]
        return []

    def get_deep_scrape_export_data_for_company(self, company_id):
        if company_id == 1:
            return [
                {
                    "phone": "0291.3957555",
                    "email": "sales@example.com",
                    "address": "Address A Longer, Bình Dương",
                    "website": "https://a.vn",
                    "source_url": "https://official-a.vn/contact",
                    "search_url": "https://official-a.vn/contact",
                    "scrape_status": "success",
                    "scrape_source_type": "official_website",
                    "filter_source_type": "official_website",
                },
                {
                    "phone": "0911.892.879",
                    "email": "",
                    "address": "",
                    "website": "",
                    "source_url": "https://official-a.vn/about",
                    "search_url": "https://official-a.vn/about",
                    "scrape_status": "success",
                    "scrape_source_type": "official_website",
                    "filter_source_type": "official_website",
                },
                {
                    "phone": "",
                    "email": "",
                    "address": "",
                    "website": "",
                    "source_url": "https://official-a.vn/jobs",
                    "search_url": "https://official-a.vn/jobs",
                    "scrape_status": "success",
                    "scrape_source_type": "official_website",
                    "filter_source_type": "official_website",
                },
                {
                    "phone": "+1 555 123 4567",
                    "email": "foreign@example.com",
                    "address": "USA",
                    "website": "https://foreign.example",
                    "source_url": "https://abc.us/contact",
                    "search_url": "https://abc.us/contact",
                    "scrape_status": "success",
                    "scrape_source_type": "official_website",
                    "filter_source_type": "official_website",
                },
            ]
        return []

    def get_pipeline_time_for_company(self, company_id):
        return {"started_at": "2026-06-22 08:00:00", "finished_at": "2026-06-22 09:00:00"}

    def fetch_one(self, *args, **kwargs):
        return {"cnt": 3}


def test_consolidated_report_preserves_source_rows_and_replaces_summary(tmp_path):
    output = tmp_path / "report.xlsx"
    ExcelWriter().write_consolidated_report(FakeDb(), str(output))

    wb = openpyxl.load_workbook(output)
    detail = wb["Detail"]
    summary = wb["Summary"]

    headers = [cell.value for cell in detail[1]]
    assert "Business Status" in headers
    assert "Business Status Category" in headers
    assert "Business Status Source URL" in headers
    assert "Province" in headers

    phone_col = headers.index("Phone") + 1
    province_col = headers.index("Province") + 1
    phones = [detail.cell(row=i, column=phone_col).value for i in range(2, detail.max_row + 1)]
    provinces = [detail.cell(row=i, column=province_col).value for i in range(2, detail.max_row + 1)]
    assert phones.count("02913957555") == 2
    assert "0911892879" in phones
    assert "15551234567" not in phones
    assert "binh duong" in provinces

    metrics = {summary.cell(row=i, column=1).value: summary.cell(row=i, column=2).value for i in range(1, 8)}
    assert metrics["Total Companies"] == 2
    assert metrics["Total Phone Numbers"] == 3
    assert metrics["Phone Coverage Rate"] == "50.00%"
    assert metrics["Companies Skipped By Inactive Status"] == 1
    assert metrics["Foreign URLs Skipped Before Scrape"] == 3
    assert metrics["Scraped Pages Excluded From Report"] == 1

    assert "Foreign Exclusions" in wb.sheetnames
    foreign = wb["Foreign Exclusions"]
    foreign_rows = list(foreign.iter_rows(values_only=True))
    assert any(row[2] == "abc.us" and row[3] == "foreign_domain_scraped" for row in foreign_rows)

    rows = list(summary.iter_rows(values_only=True))
    assert any(row[0] == "official-a.vn" and row[1] == 2 and row[2] == 3 for row in rows)
