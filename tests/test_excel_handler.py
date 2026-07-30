import json
import os
import openpyxl
import pytest
import logging
from src.database import DatabaseManager
from src.excel_handler import ExcelReader, ExcelWriter

logging.basicConfig(level=logging.INFO)

# Run tests with: pytest tests/test_excel_handler.py -v

def create_fake_excel(file_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Headers
    ws.append(["STT", " COMPANY NAME (English)", "Tax Code", "ADDRESS (English)"])
    
    # Data rows
    ws.append([1, "ABC Software Solutions Co., Ltd", "0123456789", "Hanoi"])
    ws.append([2, "XYZ Trading", None, "HCMC"])
    ws.append([3, "DEF Manufacturing TNHH", "9876543210", "Da Nang"])
    
    # Empty row
    ws.append([None, None, None, None])
    
    # Another data row with tax code
    ws.append([4, "GHI Services", "1112223334", "Can Tho"])
    
    wb.save(file_path)

def test_excel_reader(tmp_path):
    fake_excel_path = tmp_path / "fake_input.xlsx"
    create_fake_excel(str(fake_excel_path))
    
    reader = ExcelReader()
    companies = reader.read_company_list(str(fake_excel_path))
    
    assert len(companies) == 4
    
    # Verify first company
    assert companies[0]["name"] == "ABC Software Solutions Co., Ltd"
    assert companies[0]["tax_code"] == "0123456789"
    
    # Verify company without tax code
    assert companies[1]["name"] == "XYZ Trading"
    assert companies[1]["tax_code"] is None
    
    # Verify last company
    assert companies[3]["name"] == "GHI Services"
    assert companies[3]["tax_code"] == "1112223334"

def test_excel_reader_real_file():
    # If the file exists, test it
    real_file = "PIC 수집 시도_글투실_20260409.xlsx"
    if os.path.exists(real_file):
        reader = ExcelReader()
        companies = reader.read_company_list(real_file)
        assert len(companies) > 0

def test_excel_writer(tmp_path):
    output_excel_path = tmp_path / "fake_output.xlsx"
    
    results = [
        {
            "name": "Công ty A",
            "tax_code": "0123456789",
            "sources": [
                {
                    "source": "masothue",
                    "address": "123 Nguyễn Huệ",
                    "phone": "—",
                    "email": "—",
                    "website": "—",
                    "fax": "—",
                    "rep": "Nguyễn Văn A",
                    "date": "2026-04-13"
                },
                {
                    "source": "website",
                    "address": "123 Nguyễn Huệ",
                    "phone": "028-1234",
                    "email": "info@a.com",
                    "website": "a.com",
                    "fax": "—",
                    "rep": "Nguyễn Văn A",
                    "date": "2026-04-13"
                }
            ]
        },
        {
            "name": "Công ty B",
            "tax_code": "9876543210",
            "sources": [
                {
                    "source": "topcv",
                    "address": "Q1, HCM",
                    "phone": "0901-234",
                    "email": "hr@b.com",
                    "website": "b.com",
                    "fax": "—",
                    "rep": "Trần Thị B",
                    "date": "2026-04-13"
                }
            ]
        }
    ]
    
    writer = ExcelWriter()
    writer.write_results(str(output_excel_path), results)
    
    assert os.path.exists(str(output_excel_path))
    
    # Read back to verify
    wb = openpyxl.load_workbook(str(output_excel_path))
    ws = wb.active
    
    assert ws.title == "Kết quả thu thập"
    
    # Check headers
    headers = [cell.value for cell in ws[1]]
    expected_headers = [
        "STT", "Tên công ty", "Mã số thuế", "Nguồn", "Địa chỉ", 
        "SĐT", "Email", "Website", "Fax", "Người đại diện", "Ngày thu thập"
    ]
    assert headers[:len(expected_headers)] == expected_headers
    
    # Check data (row 2 -> Công ty A, masothue)
    row2 = [cell.value for cell in ws[2]]
    assert row2[1] == "Công ty A"
    assert row2[2] == "0123456789"
    assert row2[3] == "masothue"
    
    # Check data (row 3 -> Công ty A, website) - Note that the prompt said tax code could be blank 
    # but the current implementation repeats the tax_code. It's fine to repeat for easier filtering.
    row3 = [cell.value for cell in ws[3]]
    assert row3[1] == "Công ty A"
    assert row3[3] == "website"
    
    # Check data (row 4 -> Công ty B)
    row4 = [cell.value for cell in ws[4]]
    assert row4[1] == "Công ty B"
    assert row4[3] == "topcv"


def test_consolidated_report_plain_text_domains_steps_and_phone_normalization(tmp_path):
    db = DatabaseManager(str(tmp_path / "export.db"))
    db.init_db()

    gemini_grounding_id = db.insert_company(
        "Gemini Grounding Corp",
        vietnamese_name="Cong ty Gemini Grounding",
        tax_code="0123456789",
        status="done",
    )
    db.execute_query(
        """
        INSERT INTO gemini_quick_results
            (company_id, address, phone, email, website, tax_code, confidence,
             sources_json, grounding_sources_json, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            gemini_grounding_id,
            "Grounding Address",
            "+842812345678",
            "grounding@example.com",
            "https://grounding.example.com",
            "0123456789",
            0.95,
            json.dumps(["https://wrong.example.vn/company"]),
            json.dumps([
                "https://www.masothue.com/company/gemini-grounding",
                "https://other.example.vn/source",
            ]),
            "2026-04-01 10:11:12",
        ),
    )

    gemini_fallback_id = db.insert_company("Gemini Fallback Corp", tax_code="0000000001")
    db.execute_query(
        """
        INSERT INTO gemini_quick_results
            (company_id, address, phone, sources_json, grounding_sources_json, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (
            gemini_fallback_id,
            "Fallback Address",
            "84901234567",
            json.dumps([{"url": "https://www.fallback.vn/contact"}]),
            json.dumps([]),
            "2026-04-02T03:04:05+07:00",
        ),
    )

    gemini_no_source_id = db.insert_company("Gemini No Source Corp")
    db.execute_query(
        """
        INSERT INTO gemini_quick_results
            (company_id, address, phone, sources_json, grounding_sources_json, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (
            gemini_no_source_id,
            "No Source Address",
            "0909 888 777",
            json.dumps([]),
            json.dumps([]),
            "2026-04-03 00:00:00",
        ),
    )

    deep_id = db.insert_company("Deep Scrape Corp", tax_code="0987654321")
    search_id = db.insert_search_result(
        deep_id,
        "deep scrape query",
        "organic",
        1,
        "https://search-result.example.vn/company",
        "Deep Result",
        "Snippet",
    )
    db.execute_query(
        "UPDATE search_results SET created_at = ? WHERE id = ?",
        ("2026-04-04 01:02:03", search_id),
    )
    filtered_id = db.insert_filtered_link(
        search_id,
        deep_id,
        "https://www.deep.vn/contact",
        "official_website",
    )
    scraped_id = db.insert_scraped_page(
        filtered_id,
        deep_id,
        "https://www.deep.vn/contact",
        "official_website",
        "phone 0901-234-567",
        20,
        "success",
    )
    db.insert_extracted_contact(
        deep_id,
        scraped_id,
        "official_website",
        "https://www.deep.vn/contact",
        "Deep Address",
        "0901-234-567",
        "deep@example.vn",
        "https://www.deep.vn",
        None,
        None,
        "{}",
        0.9,
    )


    gemini_placeholder_id = db.insert_company("Gemini Placeholder Corp", tax_code="0000000002")
    db.execute_query(
        """
        INSERT INTO gemini_quick_results
            (company_id, address, phone, email, sources_json, grounding_sources_json, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            gemini_placeholder_id,
            "Placeholder Address",
            "N/A",
            "N/A",
            json.dumps(["https://placeholder.example.vn/contact"]),
            json.dumps([]),
            "2026-04-05 09:08:07",
        ),
    )

    deep_placeholder_id = db.insert_company("Deep Placeholder Corp", tax_code="0000000003")
    placeholder_search_id = db.insert_search_result(
        deep_placeholder_id,
        "placeholder query",
        "organic",
        1,
        "https://placeholder-search.example.vn/company",
        "Placeholder Result",
        "Snippet",
    )
    db.execute_query(
        "UPDATE search_results SET created_at = ? WHERE id = ?",
        ("2026-04-06 01:02:03", placeholder_search_id),
    )
    placeholder_filtered_id = db.insert_filtered_link(
        placeholder_search_id,
        deep_placeholder_id,
        "https://www.placeholder.vn/contact",
        "official_website",
    )
    placeholder_scraped_id = db.insert_scraped_page(
        placeholder_filtered_id,
        deep_placeholder_id,
        "https://www.placeholder.vn/contact",
        "official_website",
        "placeholder markdown",
        20,
        "success",
    )
    db.insert_extracted_contact(
        deep_placeholder_id,
        placeholder_scraped_id,
        "official_website",
        "https://www.placeholder.vn/contact",
        "Placeholder Deep Address",
        "A / N / -",
        "N/A",
        None,
        None,
        None,
        "{}",
        0.9,
    )

    output_path = tmp_path / "consolidated.xlsx"
    ExcelWriter().write_consolidated_report(
        db,
        str(output_path),
        company_ids=[
            gemini_grounding_id,
            gemini_fallback_id,
            gemini_no_source_id,
            deep_id,
            gemini_placeholder_id,
            deep_placeholder_id,
        ],
    )

    wb = openpyxl.load_workbook(output_path)
    detail = wb["Detail"]
    summary = wb["Summary"]

    headers = [cell.value for cell in detail[1]]
    assert headers == [
        "Company Name", "Vietnamese Name", "Tax Code", "Result Date",
        "Business Status", "Business Status Category", "Business Status Source URL",
        "Address", "Province", "Phone", "Source Domain", "Source URL", "Source Step",
        "Email", "Website", "Status",
    ]
    assert "Start Time" not in headers
    assert "End Time" not in headers

    for row in detail.iter_rows():
        for cell in row:
            if cell.value is not None:
                assert cell.number_format == "@"

    rows = {
        detail.cell(row=row_idx, column=1).value: [cell.value for cell in detail[row_idx]]
        for row_idx in range(2, detail.max_row + 1)
    }

    grounding = rows["Gemini Grounding Corp"]
    assert grounding[2] == "0123456789"
    assert grounding[3] == "2026-04-01"
    assert grounding[8] == "—"
    assert grounding[9] == "02812345678"
    assert grounding[10] == "masothue.com"
    assert grounding[11] == "https://www.masothue.com/company/gemini-grounding"
    assert grounding[12] == "Gemini Quick"

    fallback = rows["Gemini Fallback Corp"]
    assert fallback[9] == "0901234567"
    assert fallback[10] == "fallback.vn"
    assert fallback[11] == "https://www.fallback.vn/contact"
    assert fallback[12] == "Gemini Quick"

    no_source = rows["Gemini No Source Corp"]
    assert no_source[10] == "—"
    assert no_source[11] == "[old or unreliable source]"
    assert no_source[12] == "Gemini Quick"

    deep = rows["Deep Scrape Corp"]
    assert deep[3] == "2026-04-04"
    assert deep[9] == "0901234567"
    assert deep[10] == "deep.vn"
    assert deep[11] == "https://www.deep.vn/contact"
    assert deep[12] == "Deep Scrape"


    gemini_placeholder = rows["Gemini Placeholder Corp"]
    assert gemini_placeholder[9] == "—"
    assert gemini_placeholder[10] == "placeholder.example.vn"
    assert gemini_placeholder[11] == "https://placeholder.example.vn/contact"
    assert gemini_placeholder[12] == "Gemini Quick"
    assert gemini_placeholder[13] == "—"

    deep_placeholder = rows["Deep Placeholder Corp"]
    assert deep_placeholder[9] == "—"
    assert deep_placeholder[10] == "placeholder.vn"
    assert deep_placeholder[11] == "https://www.placeholder.vn/contact"
    assert deep_placeholder[12] == "Deep Scrape"
    assert deep_placeholder[13] == "—"
