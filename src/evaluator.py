import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Dict

logger = logging.getLogger(__name__)

class QualityEvaluator:
    def __init__(self, db):
        self.db = db

    def evaluate_extraction_quality(self, company_id: int) -> dict:
        company = self.db.get_company(company_id)
        if not company:
            return {}

        company_name = company.get("original_name", "")
        contacts = self.db.get_extracted_contacts_for_company(company_id)
        
        scraped_pages = self.db.get_scraped_pages_for_company(company_id)
        scraped_sources = set(page["source_type"] for page in scraped_pages if page.get("scrape_status") == "success")
        
        if not contacts:
            return {
                "company_id": company_id,
                "company_name": company_name,
                "quality_grade": "no_data",
                "total_sources": 0,
                "scraped_sources_count": len(scraped_sources),
                "fields_found": [],
                "fields_missing": ["address", "phone", "email", "website", "fax", "representative"],
                "cross_source_consistency": {},
                "avg_confidence": 0.0,
                "issues": ["No extracted contacts found"]
            }

        total_sources = len(contacts)
        fields = ["address", "phone", "email", "website", "fax", "representative"]
        found_fields = set()
        confidence_sum = 0
        
        source_data = {}

        for contact in contacts:
            confidence_sum += contact.get("confidence_score") or 0
            for field in fields:
                val = contact.get(field)
                if val and str(val).strip() and str(val).lower() != "none" and str(val).lower() != "null":
                    found_fields.add(field)
            
            source_type = contact.get("source_type")
            source_data[source_type] = {
                "address": str(contact.get("address", "")).lower().strip() if contact.get("address") else "",
                "phone": str(contact.get("phone", "")).lower().strip() if contact.get("phone") else ""
            }

        missing_fields = [f for f in fields if f not in found_fields]
        avg_confidence = confidence_sum / total_sources if total_sources > 0 else 0

        # Quality Grade Evaluator
        if len(found_fields) >= 3 and total_sources >= 2:
            grade = "excellent"
        elif len(found_fields) >= 2:
            grade = "good"
        elif len(found_fields) == 1:
            grade = "partial"
        else:
            grade = "no_data"

        # Cross source consistency (simple substring match check for populated ones)
        consistency = {}
        issues = []
        if total_sources >= 2:
            addresses = [data["address"] for data in source_data.values() if data["address"]]
            phones = [data["phone"] for data in source_data.values() if data["phone"]]

            if len(addresses) >= 2:
                # check if any two differ significantly
                consistency["address_match"] = len(set(addresses)) == 1
                if not consistency["address_match"]:
                    issues.append("Address mismatch between sources")
            
            if len(phones) >= 2:
                consistency["phone_match"] = len(set(phones)) == 1
                if not consistency["phone_match"]:
                    issues.append("Phone mismatch between sources")

        return {
            "company_id": company_id,
            "company_name": company_name,
            "quality_grade": grade,
            "total_sources": total_sources,
            "scraped_sources_count": len(scraped_sources),
            "fields_found": list(found_fields),
            "fields_missing": missing_fields,
            "cross_source_consistency": consistency,
            "avg_confidence": round(avg_confidence, 2),
            "issues": issues
        }

    def evaluate_batch(self, company_ids: List[int] = None) -> dict:
        if not company_ids:
            companies = self.db.get_all_companies()
            company_ids = [c["id"] for c in companies]

        results = []
        grade_dist = {"excellent": 0, "good": 0, "partial": 0, "no_data": 0}
        all_issues = []
        total_score = 0

        for cid in company_ids:
            res = self.evaluate_extraction_quality(cid)
            if res:
                results.append(res)
                grade = res.get("quality_grade", "no_data")
                grade_dist[grade] = grade_dist.get(grade, 0) + 1
                all_issues.extend(res.get("issues", []))
                
                # Assign a score
                if grade == "excellent": total_score += 100
                elif grade == "good": total_score += 70
                elif grade == "partial": total_score += 30

        total_companies = len(results)
        overall_quality_score = total_score / total_companies if total_companies > 0 else 0

        # top issues
        from collections import Counter
        common_issues = [issue for issue, count in Counter(all_issues).most_common(5)]

        recommendations = []
        partial_no_data = grade_dist["partial"] + grade_dist["no_data"]
        excellent_good = grade_dist["excellent"] + grade_dist["good"]
        if total_companies > 0 and (partial_no_data / total_companies) > 0.3:
            recommendations.append("High percentage of partial or no_data cases. Consider reviewing Scrape results or Prompt template.")
        if "Phone mismatch between sources" in common_issues:
            recommendations.append("Phone mismatch detected. Add logic to standardize phone formats in AI Prompts.")
        if "Address mismatch between sources" in common_issues:
            recommendations.append("Address mismatch detected. Address extraction may need exact substring strictness or AI mapping to standard addresses.")
        if overall_quality_score < 60:
            recommendations.append("Overall quality is low (<60%). DO NOT proceed to Phase 4 without improvements.")
        else:
            recommendations.append("Overall quality is acceptable or good. Proceed with tuning or to Phase 4.")

        return {
            "detailed_results": results,
            "grade_distribution": grade_dist,
            "overall_quality_score": round(overall_quality_score, 2),
            "common_issues": common_issues,
            "recommendations": recommendations   
        }

    def generate_evaluation_report(self, output_path: str, eval_results: dict):
        wb = openpyxl.Workbook()
        
        # Sheet 1: Đánh giá chi tiết
        ws1 = wb.active
        ws1.title = "Đánh giá chi tiết"
        
        headers = ["ID", "Tên công ty", "Quality Grade", "Số nguồn trích xuất", 
                   "Các trường tìm thấy", "Các trường thiếu", "Độ tin cậy TB", "Vấn đề"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws1.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            ws1.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20
        
        ws1.column_dimensions['E'].width = 30
        ws1.column_dimensions['F'].width = 30
        ws1.column_dimensions['H'].width = 40

        for r_idx, detail in enumerate(eval_results.get("detailed_results", []), start=2):
            ws1.cell(row=r_idx, column=1, value=detail.get("company_id"))
            ws1.cell(row=r_idx, column=2, value=detail.get("company_name"))
            ws1.cell(row=r_idx, column=3, value=detail.get("quality_grade"))
            ws1.cell(row=r_idx, column=4, value=detail.get("total_sources"))
            ws1.cell(row=r_idx, column=5, value=", ".join(detail.get("fields_found", [])))
            ws1.cell(row=r_idx, column=6, value=", ".join(detail.get("fields_missing", [])))
            ws1.cell(row=r_idx, column=7, value=detail.get("avg_confidence", 0))
            ws1.cell(row=r_idx, column=8, value="; ".join(detail.get("issues", [])))
            
        # Sheet 2: Thống kê chất lượng
        ws2 = wb.create_sheet(title="Thống kê chất lượng")
        ws2.column_dimensions['A'].width = 40
        ws2.column_dimensions['B'].width = 20
        
        ws2.cell(row=1, column=1, value="Chỉ số").font = Font(bold=True)
        ws2.cell(row=1, column=2, value="Giá trị").font = Font(bold=True)
        
        ws2.cell(row=2, column=1, value="Điểm chất lượng tổng quát (0-100)")
        ws2.cell(row=2, column=2, value=eval_results.get("overall_quality_score", 0))
        
        ws2.cell(row=4, column=1, value="Phân bổ phân loại (Grade)").font = Font(bold=True)
        r = 5
        for grade, count in eval_results.get("grade_distribution", {}).items():
            ws2.cell(row=r, column=1, value=grade)
            ws2.cell(row=r, column=2, value=count)
            r += 1

        # Sheet 3: Vấn đề cần xử lý
        ws3 = wb.create_sheet(title="Vấn đề cần xử lý")
        ws3.column_dimensions['A'].width = 50
        ws3.column_dimensions['B'].width = 60
        
        ws3.cell(row=1, column=1, value="Vấn đề phổ biến").font = Font(bold=True)
        for r_idx, issue in enumerate(eval_results.get("common_issues", []), start=2):
            ws3.cell(row=r_idx, column=1, value=issue)
            
        ws3.cell(row=1, column=2, value="Đề xuất cải thiện (Recommendations)").font = Font(bold=True)
        for r_idx, rec in enumerate(eval_results.get("recommendations", []), start=2):
            ws3.cell(row=r_idx, column=2, value=rec)

        wb.save(output_path)
