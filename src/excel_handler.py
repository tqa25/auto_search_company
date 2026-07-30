import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
import logging
from typing import List, Dict, Optional

from src.company_matcher import extract_province_from_address

logger = logging.getLogger(__name__)

class ExcelReader:
    """Reads input Excel files to extract company names and tax codes."""
    
    def __init__(self):
        self.name_keywords = ["company name (english)", "company name", "english name", "tên công ty", "name"]
        self.tax_code_keywords = ["tax code", "mã số thuế", "mst"]

    def _find_columns(self, sheet) -> tuple[Optional[int], Optional[int]]:
        """Finds the column indices for company name and tax code by scanning the first few rows."""
        name_col = None
        tax_col = None
        
        # Scan first 5 rows to find headers
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=5, values_only=True)):
            for col_idx, cell_value in enumerate(row):
                if not isinstance(cell_value, str):
                    continue
                cell_lower = cell_value.strip().lower()
                
                # Check for name keywords if not found yet
                if name_col is None:
                    if any(kw in cell_lower for kw in self.name_keywords):
                        name_col = col_idx
                
                # Check for tax code keywords if not found yet
                if tax_col is None:
                    if any(kw in cell_lower for kw in self.tax_code_keywords):
                        tax_col = col_idx
            
            # If both columns are found, stop scanning
            if name_col is not None and tax_col is not None:
                break
                
        return name_col, tax_col

    def read_company_list(self, file_path: str) -> List[Dict]:
        """Reads the Excel file and extracts company list."""
        logger.info(f"Reading Excel file: {file_path}")
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active
        except Exception as e:
            logger.error(f"Failed to load Excel file {file_path}: {e}")
            raise
            
        name_col, tax_col = self._find_columns(sheet)
        
        if name_col is None:
            logger.warning("Could not find company name column in the file. Trying to fallback to index 1 if it exists.")
            # Based on the sample output, company name (English) is often column index 1
            name_col = 1
        
        companies = []
        empty_rows = 0
        tax_code_count = 0
        
        # We start looking for data from the row after the header. Assuming row 3 based on sample.
        # But to be safe, we'll iterate through all rows and ignore rows without a valid string name.
        has_found_header = False
        
        for row in sheet.iter_rows(values_only=True):
            if not row or len(row) <= max(name_col or 0, tax_col or 0):
                empty_rows += 1
                continue
                
            name_val = row[name_col] if name_col is not None else None
            
            # Skip empty names or header row (heuristic: matches keyword)
            if not isinstance(name_val, str) or not name_val.strip():
                empty_rows += 1
                continue
                
            name_str = name_val.strip()
            
            # Check if this is the header row we found earlier
            if any(kw in name_str.lower() for kw in self.name_keywords):
                has_found_header = True
                continue
                
            # It's a company name
            tax_val = row[tax_col] if tax_col is not None and tax_col < len(row) else None
            tax_str = str(tax_val).strip() if tax_val is not None and str(tax_val).strip() and str(tax_val).lower() != 'none' else None
            
            companies.append({
                "name": name_str,
                "tax_code": tax_str
            })
            
            if tax_str:
                tax_code_count += 1
                
        logger.info(f"Excel read complete. Total companies: {len(companies)}, With tax code: {tax_code_count}, Empty/skipped rows: {empty_rows}")
        return companies

class ExcelWriter:
    """Writes output data to standard Excel format."""
    
    def __init__(self):
        self.headers = [
            "STT", "Tên công ty", "Mã số thuế", "Nguồn", "Địa chỉ", 
            "SĐT", "Email", "Website", "Fax", "Người đại diện", "Ngày thu thập"
        ]
        
    def write_results(self, output_path: str, results: List[Dict]):
        """
        Writes results to Excel file.
        `results` format expected:
        [
            {
                "name": "ABC Corp",
                "tax_code": "0123456789",
                "sources": [
                    {
                        "source": "masothue",
                        "address": "123 Nguyến Huệ",
                        "phone": "028",
                        "email": "info",
                        "website": "",
                        "fax": "",
                        "rep": "",
                        "date": "2026-04-13"
                    },
                    ...
                ]
            },
            ...
        ]
        """
        logger.info(f"Writing results to Excel file: {output_path}")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kết quả thu thập"
        
        # Write headers
        for col_idx, header in enumerate(self.headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Set default column width
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20
            
        # Freeze top row
        ws.freeze_panes = "A2"
        
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        row_idx = 2
        stt = 1
        
        for company in results:
            company_name = company.get("name", "")
            tax_code = company.get("tax_code", "")
            
            sources = company.get("sources", [])
            if not sources:
                # Still output a row even if no sources found
                sources = [{}]
                
            for i, source_data in enumerate(sources):
                # Display tax_code only on the first row of a company, use "—" otherwise
                display_tax_code = tax_code if i == 0 else "—"
                if not display_tax_code:
                    display_tax_code = "—"
                    
                row_data = [
                    stt,
                    company_name,
                    display_tax_code,
                    source_data.get("source", "—") or "—",
                    source_data.get("address", "—") or "—",
                    source_data.get("phone", "—") or "—",
                    source_data.get("email", "—") or "—",
                    source_data.get("website", "—") or "—",
                    source_data.get("fax", "—") or "—",
                    source_data.get("rep", "—") or "—",
                    source_data.get("date", "—") or "—"
                ]
                
                for col_idx, val in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = thin_border
                    # Only wrap text for long fields like address
                    if isinstance(val, str) and len(val) > 30:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                    else:
                        cell.alignment = Alignment(vertical="top")
                        
                row_idx += 1
            stt += 1
            
        try:
            wb.save(output_path)
            logger.info(f"Successfully saved to {output_path}")
        except Exception as e:
            logger.error(f"Failed to save to {output_path}: {e}")
            raise

    def write_final_report(self, output_path: str, aggregated_data: List[Dict], summary_stats: Dict):
        """
        Writes final report to Excel file with two sheets: details and summary stats.
        """
        logger.info(f"Writing final report to Excel file: {output_path}")
        
        wb = openpyxl.Workbook()
        ws_details = wb.active
        ws_details.title = "Kết quả thu thập"
        
        # Write headers for details sheet
        custom_headers = self.headers[:-1] + ["Độ tin cậy", "Trạng thái Pipeline", "Ngày thu thập"]
        for col_idx, header in enumerate(custom_headers, start=1):
            cell = ws_details.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws_details.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20
            
            if header == "Trạng thái Pipeline":
                ws_details.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 40
            
        ws_details.freeze_panes = "A2"
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        row_idx = 2
        stt = 1
        
        for company in aggregated_data:
            company_name = company.get("company_name", "")
            tax_code = company.get("tax_code", "")
            has_data = company.get("has_data", False)
            collection_date = company.get("collection_date", "")
            
            if not has_data:
                pipeline_status = company.get("pipeline_status_summary", "—")
                row_data = [
                    stt, company_name, tax_code, "(không tìm thấy)",
                    "—", "—", "—", "—", "—", "—", "—", pipeline_status, collection_date
                ]
                for col_idx, val in enumerate(row_data, start=1):
                    cell = ws_details.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = thin_border
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                row_idx += 1
            else:
                sources = company.get("sources", [])
                for i, source in enumerate(sources):
                    display_tax_code = tax_code if i == 0 else ""
                    display_company_name = company_name if i == 0 else ""
                    confidence = source.get("confidence")
                    pipeline_status = company.get("pipeline_status_summary", "") if i == 0 else ""
                    
                    row_data = [
                        stt if i == 0 else "",
                        display_company_name,
                        display_tax_code,
                        source.get("source_url", "—") or "—",
                        source.get("address", "—") or "—",
                        source.get("phone", "—") or "—",
                        source.get("email", "—") or "—",
                        source.get("website", "—") or "—",
                        source.get("fax", "—") or "—",
                        source.get("representative", "—") or "—",
                        confidence if confidence is not None else "—",
                        pipeline_status,
                        collection_date if i == 0 else ""
                    ]
                    
                    for col_idx, val in enumerate(row_data, start=1):
                        cell = ws_details.cell(row=row_idx, column=col_idx, value=val)
                        cell.border = thin_border
                        if isinstance(val, str) and len(val) > 30:
                            cell.alignment = Alignment(wrap_text=True, vertical="top")
                        else:
                            cell.alignment = Alignment(vertical="top")
                            
                        # Conditional formatting for confidence column (11)
                        if col_idx == 11 and isinstance(val, (int, float)):
                            if val >= 0.8:
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                                cell.font = Font(color="006100")
                            elif val >= 0.5:
                                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                                cell.font = Font(color="9C6500")
                            else:
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                                cell.font = Font(color="9C0006")
                                
                        # Conditional formatting for pipeline status column (12)
                        if col_idx == 12 and isinstance(val, str):
                            if "❌" in val:
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                                cell.font = Font(color="9C0006")
                            elif "⏭️" in val:
                                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                                cell.font = Font(color="9C6500")
                            elif "✅" in val:
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                                cell.font = Font(color="006100")
                    row_idx += 1
            stt += 1
            
        # Sheet 2: Summary Stats
        ws_summary = wb.create_sheet(title="Thống kê tổng quát")
        
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 20
        
        cell = ws_summary.cell(row=1, column=1, value="Chỉ số")
        cell.font = Font(bold=True)
        cell = ws_summary.cell(row=1, column=2, value="Giá trị")
        cell.font = Font(bold=True)
        
        summary_rows = [
            ("Tổng số công ty", summary_stats.get("total_companies", 0)),
            ("Số công ty có dữ liệu", summary_stats.get("companies_with_data", 0)),
            ("Số công ty không có dữ liệu", summary_stats.get("companies_no_data", 0)),
            ("Tỷ lệ coverage (%)", round(summary_stats.get("coverage_rate", 0), 2)),
            ("Số nguồn trung bình / công ty", round(summary_stats.get("avg_sources_per_company", 0), 2)),
            ("Độ tin cậy trung bình", round(summary_stats.get("avg_confidence", 0), 2)),
        ]
        
        r_idx = 2
        for key, val in summary_rows:
            ws_summary.cell(row=r_idx, column=1, value=key)
            ws_summary.cell(row=r_idx, column=2, value=val)
            r_idx += 1
            
        r_idx += 1
        ws_summary.cell(row=r_idx, column=1, value="Độ phủ theo trường (%)").font = Font(bold=True)
        r_idx += 1
        for field, pct in summary_stats.get("field_coverage", {}).items():
            ws_summary.cell(row=r_idx, column=1, value=field)
            ws_summary.cell(row=r_idx, column=2, value=f"{round(pct, 2)}%")
            r_idx += 1
            
        r_idx += 1
        ws_summary.cell(row=r_idx, column=1, value="Phân bổ theo nguồn (%)").font = Font(bold=True)
        r_idx += 1
        for source, pct in summary_stats.get("source_distribution", {}).items():
            ws_summary.cell(row=r_idx, column=1, value=source)
            ws_summary.cell(row=r_idx, column=2, value=f"{round(pct, 2)}%")
            r_idx += 1
            
        r_idx += 2
        ws_summary.cell(row=r_idx, column=1, value="Danh sách công ty không tìm thấy").font = Font(bold=True)
        r_idx += 1
        for company in aggregated_data:
            if not company.get("has_data", False):
                ws_summary.cell(row=r_idx, column=1, value=company.get("company_name", ""))
                ws_summary.cell(row=r_idx, column=2, value=company.get("tax_code", ""))
                r_idx += 1
                
        try:
            wb.save(output_path)
            logger.info(f"Successfully saved final report to {output_path}")
        except Exception as e:
            logger.error(f"Failed to save final report to {output_path}: {e}")
            raise

    def write_consolidated_report(self, db, output_path: str, company_ids: List[int] = None):
        """
        Writes a consolidated Excel report with one row per unique company phone.
        Tab 1: "Summary"
        Tab 2: "Detail"
        """
        import json
        import os
        import re
        import urllib.parse
        from collections import defaultdict

        logger.info(f"Writing consolidated Excel report to {output_path}")

        all_companies = db.get_all_companies()
        if company_ids:
            companies = [c for c in all_companies if c['id'] in company_ids]
        else:
            companies = all_companies

        wb = openpyxl.Workbook()

        ws_summary = wb.active
        ws_summary.title = "Summary"

        ws1 = wb.create_sheet(title="Detail")

        def normalize_phone(phone_val):
            if phone_val is None:
                return ""
            raw = str(phone_val).strip()
            if not raw:
                return ""

            has_plus = raw.lstrip().startswith("+")
            digits = re.sub(r"\D", "", raw)
            if not digits:
                return raw

            if digits.startswith("0084") and len(digits) > 4:
                return "0" + digits[4:]
            if has_plus and digits.startswith("84") and len(digits) > 2:
                return "0" + digits[2:]
            if digits.startswith("84") and len(digits) in (11, 12):
                return "0" + digits[2:]
            return digits

        def raw_phone_parts(phone_str):
            if not phone_str:
                return []
            phone_str = str(phone_str).strip()
            collapsed = re.sub(r"[\s/|,;._-]+", "", phone_str).lower()
            if collapsed in {"", "na", "none", "null", "khongcongkhai"}:
                return []
            for sep in [',', ';', '|', '/', '\n']:
                phone_str = phone_str.replace(sep, ',')
            return [p.strip() for p in phone_str.split(',') if p and p.strip()]

        def clean_phone(phone_str):
            parts = []
            for p in raw_phone_parts(phone_str):
                cleaned = normalize_phone(p)
                cleaned_compact = re.sub(r"\s+", "", cleaned).lower()
                if cleaned_compact in {"", "na", "none", "null", "khongcongkhai"}:
                    continue
                if cleaned in {"—", "-", "Không công khai", "không công khai"}:
                    continue
                if len(cleaned) <= 1 and not any(ch.isdigit() for ch in cleaned):
                    continue
                parts.append(cleaned)
            return parts

        def classify_phone_parts(phone_str):
            vn_phones = []
            foreign_phones = []
            for raw in raw_phone_parts(phone_str):
                cleaned = normalize_phone(raw)
                digits = re.sub(r"\D", "", str(raw))
                raw_text = str(raw).strip()
                if not cleaned or cleaned in {"—", "-"}:
                    continue
                is_vn = (
                    cleaned.startswith("0") and len(cleaned) in (10, 11)
                ) or raw_text.startswith("+84") or digits.startswith("0084") or digits.startswith("84")
                is_foreign = (raw_text.startswith("+") and not raw_text.startswith("+84")) or (
                    digits.startswith("00") and not digits.startswith("0084")
                )
                if is_vn:
                    vn_phones.append(cleaned)
                elif is_foreign:
                    foreign_phones.append(cleaned)
            return vn_phones, foreign_phones

        foreign_tld_suffixes = (
            ".jp", ".kr", ".cn", ".sg", ".de", ".uk", ".co.uk", ".fr", ".it",
            ".es", ".nl", ".au", ".ca", ".us", ".tw", ".hk", ".my", ".th",
            ".id", ".in", ".br", ".mx", ".ru", ".pl", ".se", ".ch",
        )

        def is_foreign_domain(domain):
            domain = str(domain or "").lower()
            return any(domain == suffix.lstrip(".") or domain.endswith(suffix) for suffix in foreign_tld_suffixes)

        def clean_email(email_str):
            if not email_str:
                return []
            email_str = str(email_str).strip()

            collapsed = re.sub(r"[\s/|,;._-]+", "", email_str).lower()
            if collapsed in {"", "na", "none", "null", "khongcongkhai"}:
                return []

            for sep in [',', ';', '|', '/', '\n']:
                email_str = email_str.replace(sep, ',')
            parts = []
            for p in email_str.split(','):
                cleaned = p.strip().lower()
                cleaned_compact = re.sub(r"[\s._-]+", "", cleaned)
                if cleaned_compact in {"", "na", "none", "null", "khongcongkhai"}:
                    continue
                if cleaned in {"—", "-"}:
                    continue
                if len(cleaned) <= 1 and "@" not in cleaned:
                    continue
                parts.append(cleaned)
            return parts

        unreliable_source_marker = "[old or unreliable source]"

        def extract_domain(url, default="Unknown"):
            if not url or url == '—' or url == unreliable_source_marker:
                return default
            try:
                if isinstance(url, dict):
                    url = url.get("url") or url.get("uri") or ""
                url = str(url).strip()
                if not url:
                    return default
                if not url.startswith('http'):
                    url = 'http://' + url
                domain = urllib.parse.urlparse(url).netloc
                if domain.startswith("www."):
                    domain = domain[4:]
                return domain if domain else default
            except Exception:
                return default

        def format_result_date(*values):
            for val in values:
                if not val or val == '—':
                    continue
                match = re.search(r"\d{4}-\d{2}-\d{2}", str(val))
                if match:
                    return match.group(0)
            return "—"

        def load_json_list(raw):
            if not raw:
                return []
            if isinstance(raw, list):
                return raw
            try:
                parsed = json.loads(raw)
            except (TypeError, json.JSONDecodeError):
                return []
            return parsed if isinstance(parsed, list) else []

        def source_url_value(source):
            if isinstance(source, dict):
                return source.get("url") or source.get("uri") or source.get("link") or ""
            return str(source).strip() if source else ""

        def first_gemini_source_url(result):
            sources = load_json_list(result.get('grounding_sources_json'))
            if not sources:
                sources = load_json_list(result.get('sources_json'))
            for source in sources:
                url = source_url_value(source)
                if url:
                    return url
            return unreliable_source_marker

        def write_text_cell(ws, row, column, value, border=None, alignment=None):
            if value is None or value == "":
                value = "—"
            value = str(value)
            if len(value) > 200:
                value = value[:197] + "..."
            cell = ws.cell(row=row, column=column, value=value)
            cell.number_format = "@"
            if border:
                cell.border = border
            if alignment:
                cell.alignment = alignment
            return cell

        flattened_rows = []
        status_map = {
            'done': 'Done',
            'failed': 'Failed',
            'permanently_failed': 'Failed',
            'pending': 'Pending',
            'running': 'Running',
            'scraping': 'Scraping',
        }

        official_domain_stats = defaultdict(lambda: {"pages": set(), "phone_pages": set()})
        companies_with_phone = set()
        inactive_skipped_count = 0
        foreign_exclusions = []
        companies_with_foreign_only = set()
        foreign_urls_skipped_before_scrape = 0
        if hasattr(db, "fetch_one"):
            row = db.fetch_one(
                """
                SELECT COUNT(*) AS cnt
                FROM filtered_links
                WHERE should_scrape = 0
                  AND (reason LIKE 'foreign_tld_skip:%'
                       OR reason LIKE 'name_overmatch_skip:%'
                       OR reason LIKE '%weak_vietnam_identity%')
                """
            ) or {}
            foreign_urls_skipped_before_scrape = row.get("cnt", 0) or 0

        for company in companies:
            cid = company['id']
            c_name = company['original_name']
            vn_name = company.get('vietnamese_name', '') or '—'
            t_code = company.get('tax_code', '') or '—'
            db_status = company.get('status', 'pending')
            status_display = status_map.get(db_status, db_status)
            business_status = company.get('business_status') or '—'
            business_status_category = company.get('business_status_category') or '—'
            business_status_source_url = company.get('business_status_source_url') or '—'
            if business_status_category == 'inactive_stop':
                inactive_skipped_count += 1

            gemini_results = db.get_gemini_quick_results_for_company(cid)
            deep_scrape_data = db.get_deep_scrape_export_data_for_company(cid)
            time_data = db.get_pipeline_time_for_company(cid)

            fallback_result_date = format_result_date(
                time_data.get('finished_at'),
                time_data.get('started_at')
            )

            source_rows = []
            company_foreign_contacts = False

            def append_source_rows(phone_values, address, email, website, domain, source_url, source_step, result_date):
                address = address.strip() if isinstance(address, str) and address.strip() else "—"
                province = extract_province_from_address(address) if address != "—" else None
                province = province if province else "—"

                email_values = clean_email(email)
                email_value = ", ".join(email_values) if email_values else "—"
                website_value = website.strip() if isinstance(website, str) and website.strip() else "—"

                if phone_values:
                    for phone in phone_values:
                        companies_with_phone.add(cid)
                        source_rows.append([
                            c_name, vn_name, t_code, result_date,
                            business_status, business_status_category, business_status_source_url,
                            address, province, phone, domain, source_url, source_step,
                            email_value, website_value, status_display,
                        ])
                elif any(value != "—" for value in [address, email_value, website_value]):
                    source_rows.append([
                        c_name, vn_name, t_code, result_date,
                        business_status, business_status_category, business_status_source_url,
                        address, province, "—", domain, source_url, source_step,
                        email_value, website_value, status_display,
                    ])

            for gr in gemini_results:
                source_url = first_gemini_source_url(gr)
                domain = extract_domain(source_url, default="—")
                result_date = format_result_date(gr.get('created_at'), fallback_result_date)
                phones, foreign_phones = classify_phone_parts(gr.get('phone')) if gr.get('phone') else ([], [])
                if foreign_phones:
                    company_foreign_contacts = True
                    foreign_exclusions.append({
                        "company_name": c_name,
                        "tax_code": t_code,
                        "domain": domain,
                        "source_url": source_url,
                        "reason": "foreign_phone_only" if not phones else "foreign_phone_excluded",
                        "phones": ", ".join(foreign_phones),
                    })
                append_source_rows(
                    phones,
                    gr.get('address'),
                    gr.get('email'),
                    gr.get('website'),
                    domain,
                    source_url,
                    "Gemini Quick",
                    result_date,
                )

            for row in deep_scrape_data:
                source_url = row.get('source_url') or row.get('search_url') or '—'
                domain = extract_domain(source_url, default="Unknown")
                source_type = row.get('scrape_source_type') or row.get('filter_source_type')
                row_phones, foreign_phones = classify_phone_parts(row.get('phone')) if row.get('phone') else ([], [])
                foreign_domain = is_foreign_domain(domain)
                exclude_row = foreign_domain or (foreign_phones and not row_phones)
                if exclude_row:
                    company_foreign_contacts = True
                    reason = "foreign_domain_scraped" if foreign_domain else "foreign_phone_only"
                    foreign_exclusions.append({
                        "company_name": c_name,
                        "tax_code": t_code,
                        "domain": domain,
                        "source_url": source_url,
                        "reason": reason,
                        "phones": ", ".join(foreign_phones),
                    })
                    continue
                if foreign_phones:
                    company_foreign_contacts = True
                    foreign_exclusions.append({
                        "company_name": c_name,
                        "tax_code": t_code,
                        "domain": domain,
                        "source_url": source_url,
                        "reason": "foreign_phone_excluded",
                        "phones": ", ".join(foreign_phones),
                    })
                if source_type == 'official_website' and row.get('scrape_status') == 'success' and domain != 'Unknown':
                    official_domain_stats[domain]["pages"].add(source_url)
                    if row_phones:
                        official_domain_stats[domain]["phone_pages"].add(source_url)
                result_date = format_result_date(row.get('timestamp'), fallback_result_date)
                append_source_rows(
                    row_phones,
                    row.get('address'),
                    row.get('email'),
                    row.get('website'),
                    domain,
                    source_url,
                    "Deep Scrape",
                    result_date,
                )

            if not source_rows:
                if company_foreign_contacts:
                    companies_with_foreign_only.add(cid)
                flattened_rows.append([
                    c_name, vn_name, t_code, fallback_result_date,
                    business_status, business_status_category, business_status_source_url,
                    "—", "—", "—", "—", "—", "—", "—", "—", status_display,
                ])
            else:
                flattened_rows.extend(source_rows)

        headers = [
            "Company Name", "Vietnamese Name", "Tax Code", "Result Date",
            "Business Status", "Business Status Category", "Business Status Source URL",
            "Address", "Province", "Phone", "Source Domain", "Source URL", "Source Step",
            "Email", "Website", "Status",
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = write_text_cell(ws1, 1, col_idx, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws1.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

        ws1.freeze_panes = "A2"

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row_idx, row_data in enumerate(flattened_rows, start=2):
            for col_idx, val in enumerate(row_data, start=1):
                write_text_cell(
                    ws1, row_idx, col_idx, val,
                    border=thin_border,
                    alignment=Alignment(vertical="top")
                )

        ws_summary.column_dimensions['A'].width = 46
        ws_summary.column_dimensions['B'].width = 24
        ws_summary.column_dimensions['C'].width = 18
        ws_summary.column_dimensions['D'].width = 18

        write_text_cell(ws_summary, 1, 1, "Metric").font = Font(bold=True)
        write_text_cell(ws_summary, 1, 2, "Value").font = Font(bold=True)

        total_companies = len(companies)
        total_phones = len([row for row in flattened_rows if row[9] != '—'])
        phone_coverage = (len(companies_with_phone) / total_companies * 100) if total_companies else 0
        min_sample_size = 3
        summary_rows = [
            ("Total Companies", total_companies),
            ("Total Phone Numbers", total_phones),
            ("Phone Coverage Rate", f"{phone_coverage:.2f}%"),
            ("Companies Skipped By Inactive Status", inactive_skipped_count),
            ("Foreign URLs Skipped Before Scrape", foreign_urls_skipped_before_scrape),
            ("Scraped Pages Excluded From Report", len(foreign_exclusions)),
            ("Companies With Only Foreign Contact", len(companies_with_foreign_only)),
            ("Official Domains Minimum Sample Size", min_sample_size),
        ]
        for idx, (metric, value) in enumerate(summary_rows, start=2):
            ws_summary.cell(row=idx, column=1, value=metric)
            ws_summary.cell(row=idx, column=2, value=value)

        current_sum_row = len(summary_rows) + 4
        write_text_cell(
            ws_summary,
            current_sum_row,
            1,
            "Top 5 Official Websites With Best Phone Results"
        ).font = Font(bold=True)
        current_sum_row += 1
        for col_idx, header in enumerate(["Domain", "Phone Pages", "Scraped Pages", "Phone Rate"], start=1):
            write_text_cell(ws_summary, current_sum_row, col_idx, header).font = Font(bold=True)
        current_sum_row += 1

        ranked_domains = []
        for domain, stat in official_domain_stats.items():
            scraped_pages = len(stat["pages"])
            if scraped_pages < min_sample_size:
                continue
            phone_pages = len(stat["phone_pages"])
            rate = phone_pages / scraped_pages if scraped_pages else 0
            ranked_domains.append((rate, phone_pages, scraped_pages, domain))
        ranked_domains.sort(reverse=True)

        for rate, phone_pages, scraped_pages, domain in ranked_domains[:5]:
            write_text_cell(ws_summary, current_sum_row, 1, domain)
            ws_summary.cell(row=current_sum_row, column=2, value=phone_pages)
            ws_summary.cell(row=current_sum_row, column=3, value=scraped_pages)
            write_text_cell(ws_summary, current_sum_row, 4, f"{rate * 100:.2f}%")
            current_sum_row += 1

        current_sum_row += 2
        write_text_cell(ws_summary, current_sum_row, 1, "Foreign Exclusions").font = Font(bold=True)
        current_sum_row += 1
        for col_idx, header in enumerate(["Company", "Tax Code", "Domain", "Reason", "Phones", "Source URL"], start=1):
            write_text_cell(ws_summary, current_sum_row, col_idx, header).font = Font(bold=True)
        current_sum_row += 1
        for item in foreign_exclusions[:50]:
            write_text_cell(ws_summary, current_sum_row, 1, item.get("company_name"))
            write_text_cell(ws_summary, current_sum_row, 2, item.get("tax_code"))
            write_text_cell(ws_summary, current_sum_row, 3, item.get("domain"))
            write_text_cell(ws_summary, current_sum_row, 4, item.get("reason"))
            write_text_cell(ws_summary, current_sum_row, 5, item.get("phones"))
            write_text_cell(ws_summary, current_sum_row, 6, item.get("source_url"))
            current_sum_row += 1

        ws_foreign = wb.create_sheet(title="Foreign Exclusions")
        for col_idx, header in enumerate(["Company", "Tax Code", "Domain", "Reason", "Phones", "Source URL"], start=1):
            cell = write_text_cell(ws_foreign, 1, col_idx, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        for row_idx, item in enumerate(foreign_exclusions, start=2):
            write_text_cell(ws_foreign, row_idx, 1, item.get("company_name"), border=thin_border)
            write_text_cell(ws_foreign, row_idx, 2, item.get("tax_code"), border=thin_border)
            write_text_cell(ws_foreign, row_idx, 3, item.get("domain"), border=thin_border)
            write_text_cell(ws_foreign, row_idx, 4, item.get("reason"), border=thin_border)
            write_text_cell(ws_foreign, row_idx, 5, item.get("phones"), border=thin_border)
            write_text_cell(ws_foreign, row_idx, 6, item.get("source_url"), border=thin_border)
        for col_idx in range(1, 7):
            ws_foreign.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 24

        for row_idx in range(1, current_sum_row):
            for col_idx in range(1, 7):
                ws_summary.cell(row=row_idx, column=col_idx).border = thin_border

        for col in ws1.columns:
            max_len = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 50))
            ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)

        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
        wb.save(output_path)
        logger.info(f"Successfully saved consolidated report to {output_path}")
